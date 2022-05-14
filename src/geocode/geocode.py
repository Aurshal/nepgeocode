from openpyxl import load_workbook
from thefuzz import process
import json

wb = load_workbook('src/geocode/data/geocode.xlsx')
ws_district = wb["district_code"]
ws_localunit = wb["localunit_code"]
ws_pradesh = wb["pradesh_code"]
try:
    with open("src/geocode/data/dist_code.json", "r") as infile:
        dist_code_dict = json.load(infile)
except:
    dist_code_dict = {}
    for row in range(4, ws_district.max_row + 1):
        dist_name = ws_district["C" + str(row)].value
        dist_code = ws_district["E" + str(row)].value
        dist_code_dict[dist_name.title()] = dist_code
    jd = json.dumps(dist_code_dict, indent=4, ensure_ascii=False)
    with open("src/geocode/data/dist_code.json", "w") as outfile:
        outfile.write(jd)
try:
    with open("src/geocode/data/mun_code.json", "r") as infile:
        mun_code_dict = json.load(infile)
except:
    mun_code_dict = {}
    for row in range(4, ws_localunit.max_row + 1):
        mun_name = ws_localunit["C" + str(row)].value
        mun_code = ws_localunit["E" + str(row)].value
        mun_dist = ws_localunit["B" + str(row)].value
        mun_name = mun_name.split()
        mun_name.pop()
        mun_name.pop()
        mun_name = " ".join(mun_name)
        if not mun_code_dict.get(mun_name):
            mun_code_dict[mun_name] = [
                [mun_code, mun_dist],
            ]
        else:
            mun_code_dict[mun_name].append([mun_code, mun_dist])
    jd = json.dumps(mun_code_dict, indent=4, ensure_ascii=False)
    with open("src/geocode/data/mun_code.json", "w") as outfile:
        outfile.write(jd)


def get_code(query: str, d_n=None):
    dp = []
    matched_query = process.extract(
        query, list(dist_code_dict.keys()) + list(mun_code_dict.keys())
    )
    query = matched_query[0][0]
    ratio = matched_query[0][1]
    if ratio < 80:
        return None
    for q in matched_query:
        if query in q[0]:
            dp.append(q[0])
    if d_n:
        m_d = process.extract(d_n, list(dist_code_dict.keys()))
        if m_d[0][1] >= 80:
            d_n = m_d[0][0]
        else:
            return None
        for q in dp:
            for val in mun_code_dict[q]:
                if d_n in val:
                    return val[0]
    else:
        try:
            return dist_code_dict[dp[0]]
        except:
            if len(mun_code_dict[dp[0]]) == 1:
                return mun_code_dict[dp[0]][0][0]
            else:
                print("Multiple municipalities found. Please mention district.")
                return None
    return None


